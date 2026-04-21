"""
04_quality_stress.py

Runs olgadoc, calamine and openpyxl on stress.xlsx, extracts the raw text,
saves it to results/, then scores each extractor on 18 quality checks that
directly target documented calamine limitations.

Usage:  python3 04_quality_stress.py
"""
import pathlib
import re
import subprocess
import sys

ROOT = pathlib.Path(__file__).parent.parent
FIXTURES = ROOT / "fixtures"
RESULTS  = ROOT / "results"
RESULTS.mkdir(exist_ok=True)
HARNESS = ROOT / "scripts" / "02_harness.py"

EXTRACTORS = ["olgadoc", "calamine", "openpyxl"]


def extract_text(lib: str, path: pathlib.Path) -> str:
    """Reuse the harness to get deterministic text output."""
    import time as _t

    if lib == "olgadoc":
        import olgadoc
        return olgadoc.Document.open(str(path)).text()

    if lib == "calamine":
        from python_calamine import CalamineWorkbook
        wb = CalamineWorkbook.from_path(str(path))
        parts = []
        for name in wb.sheet_names:
            parts.append(f"### {name}")
            for row in wb.get_sheet_by_name(name).to_python():
                parts.append("\t".join("" if v is None else str(v) for v in row))
        return "\n".join(parts)

    if lib == "openpyxl":
        from openpyxl import load_workbook
        wb = load_workbook(str(path), data_only=True, read_only=True)
        parts = []
        for name in wb.sheetnames:
            parts.append(f"### {name}")
            for row in wb[name].iter_rows(values_only=True):
                parts.append("\t".join("" if v is None else str(v) for v in row))
        return "\n".join(parts)

    raise SystemExit(lib)


def section(text: str, sheet_name: str) -> str:
    """Isolate the block of output belonging to one sheet."""
    for mk in (f"### {sheet_name}", sheet_name):
        i = text.find(mk)
        if i >= 0:
            break
    else:
        return ""
    nxt = re.search(r"\n(?:### )?\d+-\w", text[i + len(mk):])
    j = (i + len(mk) + nxt.start()) if nxt else len(text)
    return text[i:j]


def format_applied(sheet_text: str, raw: str, formatted: str) -> str:
    """Did the extractor display the formatted value, not the raw one?"""
    has_raw = re.search(r"\b" + re.escape(raw) + r"\b", sheet_text) is not None
    has_fmt = formatted in sheet_text
    if has_fmt and not has_raw:
        return "✓"
    if has_raw:
        return "raw"
    return "?"


def main() -> None:
    path = FIXTURES / "stress.xlsx"
    outputs: dict[str, str] = {}
    for lib in EXTRACTORS:
        txt = extract_text(lib, path)
        (RESULTS / f"stress_{lib}.txt").write_text(txt, encoding="utf-8")
        outputs[lib] = txt

    tests = [
        # Each entry: (label, checker_fn: text → verdict_str)
        ("1. 0.15 + '0%'        → '15%'",
         lambda t: format_applied(section(t, "1-Formats"), "0.15", "15%")),
        ("1. 1234.5 + '#,##0.00'→ '1,234.50'",
         lambda t: format_applied(section(t, "1-Formats"), "1234.5", "1,234.50")),
        ("1. 7 + '000'          → '007'",
         lambda t: (format_applied(section(t, "1-Formats"), "7.0", "007")
                    if "7.0" in section(t, "1-Formats")
                    else format_applied(section(t, "1-Formats"), "7", "007"))),
        ("1. 0.5 + '# ?/?'      → '1/2'",
         lambda t: format_applied(section(t, "1-Formats"), "0.5", "1/2")),
        ('1. 1234.56 + \'"$"…\'    → \'$1,234.56\'',
         lambda t: format_applied(section(t, "1-Formats"), "1234.56", "$1,234.56")),
        ("1. 0.99999 + '0.00%'  → '100.00%'",
         lambda t: format_applied(section(t, "1-Formats"), "0.99999", "100.00%")),
        ("1. 1.2345e-05         → '1.23E-05'",
         lambda t: ("✓" if "1.23E-05" in section(t, "1-Formats")
                    and "1.2345e-05" not in section(t, "1-Formats") else "raw")),
        ("2. Duration 49h30m    → '49:30'",
         lambda t: "✓" if "49:30" in section(t, "2-DatesTimes") else "✗"),
        ("3. Merged region replicates across rows",
         lambda t: "✓×N" if section(t, "3-Merged").count("North America") >= 2 else "once"),
        ("4. Hyperlink URL attached to its label",
         lambda t: ("✓ near" if "example.com" in section(t, "4-Links")[
                                 section(t, "4-Links").find("Our website"):
                                 section(t, "4-Links").find("Our website") + 120]
                    else "col-B only" if "example.com" in t else "✗")),
        ("5. Cell comment 'Marlow-Chen' surfaces",
         lambda t: "✓" if "Marlow-Chen" in t else "✗"),
        ("7. Intra-cell newlines preserved",
         lambda t: ("✓" if "line1" in section(t, "7-LineBreaks")
                    and "line3" in section(t, "7-LineBreaks")
                    and "\n" in section(t, "7-LineBreaks")
                           [section(t, "7-LineBreaks").find("line1"):
                            section(t, "7-LineBreaks").find("line3")]
                    else "✗")),
        ("10. '#DIV/0!' surfaces",
         lambda t: "✓" if "#DIV/0" in t else "✗"),
        ("10. '#N/A' surfaces",
         lambda t: "✓" if "#N/A" in t else "✗"),
        ("11. Named range 'MyTotal' surfaces",
         lambda t: "✓" if "MyTotal" in t else "✗"),
        ("12. Hidden row leaks into output",
         lambda t: "leaks" if "HIDDEN_ROW_SECRET" in t else "filtered"),
        ("12. Hidden column leaks",
         lambda t: "leaks" if "HIDDEN_COL_SECRET" in t else "filtered"),
        ("13. VeryHidden sheet leaks",
         lambda t: "leaks" if "VERY_HIDDEN_CONTENT" in t else "filtered"),
        ("14. Data-validation values surface",
         lambda t: "✓" if "Pending" in t else "✗"),
    ]

    print(f"\n{'Test':<52} {'olgadoc':<11} {'calamine':<11} {'openpyxl':<11}")
    print("-" * 88)
    rows_for_csv = [("Test", "olgadoc", "calamine", "openpyxl")]
    for label, checker in tests:
        row = [label]
        for lib in EXTRACTORS:
            try:
                row.append(checker(outputs[lib]))
            except Exception as e:
                row.append(f"err:{type(e).__name__}")
        print(f"{row[0]:<52} {row[1]:<11} {row[2]:<11} {row[3]:<11}")
        rows_for_csv.append(tuple(row))

    import csv
    with (RESULTS / "quality_scorecard.csv").open("w", newline="") as fh:
        csv.writer(fh).writerows(rows_for_csv)
    print(f"\n→ wrote {RESULTS / 'quality_scorecard.csv'}")


if __name__ == "__main__":
    main()
