"""
02_harness.py

Process-isolated benchmark harness. Runs ONE extractor on ONE file and
prints a JSON result line. Invoked by 03_bench.py as a subprocess so that
each extractor starts in a cold interpreter with ONLY its own imports loaded
— no cross-contamination, no cached dynamic-lib warm-up advantage.

Usage:   python3 02_harness.py <file> <extractor>
         python3 02_harness.py invoice.pdf olgadoc
"""
import json
import sys
import time


def load(lib: str, path: str) -> str:
    """Return the extracted text as a single string."""
    if lib == "olgadoc":
        import olgadoc
        return olgadoc.Document.open(path).text()

    if lib == "calamine":
        from python_calamine import CalamineWorkbook
        wb = CalamineWorkbook.from_path(path)
        out: list[str] = []
        for name in wb.sheet_names:
            out.append(f"### {name}")
            ws = wb.get_sheet_by_name(name)
            for row in ws.to_python():
                out.append("\t".join("" if v is None else str(v) for v in row))
        return "\n".join(out)

    if lib == "openpyxl":
        from openpyxl import load_workbook
        wb = load_workbook(path, data_only=True, read_only=True)
        out = []
        for name in wb.sheetnames:
            out.append(f"### {name}")
            ws = wb[name]
            for row in ws.iter_rows(values_only=True):
                out.append("\t".join("" if v is None else str(v) for v in row))
        return "\n".join(out)

    if lib == "pdfplumber":
        import pdfplumber
        with pdfplumber.open(path) as p:
            return "\n".join((pg.extract_text() or "") for pg in p.pages)

    if lib == "pymupdf":
        import fitz
        d = fitz.open(path)
        return "\n".join(pg.get_text() for pg in d)

    if lib == "pypdf":
        from pypdf import PdfReader
        r = PdfReader(path)
        return "\n".join(pg.extract_text() or "" for pg in r.pages)

    if lib == "docx2txt":
        import docx2txt
        return docx2txt.process(path)

    if lib == "mammoth":
        import mammoth
        with open(path, "rb") as f:
            return mammoth.extract_raw_text(f).value

    if lib == "python-docx":
        from docx import Document
        d = Document(path)
        return "\n".join(p.text for p in d.paragraphs)

    if lib == "bs4":
        from bs4 import BeautifulSoup
        with open(path, encoding="utf-8") as f:
            return BeautifulSoup(f, "lxml").get_text("\n", strip=True)

    if lib == "trafilatura":
        import trafilatura
        return trafilatura.extract(open(path, encoding="utf-8").read()) or ""

    if lib == "html2text":
        import html2text
        h = html2text.HTML2Text()
        h.body_width = 0
        return h.handle(open(path, encoding="utf-8").read())

    raise SystemExit(f"unknown extractor: {lib}")


def main() -> None:
    if len(sys.argv) != 3:
        sys.exit("usage: 02_harness.py <file> <extractor>")
    path, lib = sys.argv[1], sys.argv[2]

    # One warm-up call (loads shared libs, JITs, fills page caches).
    try:
        txt_warm = load(lib, path)
    except Exception as e:
        print(json.dumps({"error": f"{type(e).__name__}: {e}"}))
        return

    # Ten timed calls, keep the best (min) and the median.
    runs_ms: list[float] = []
    for _ in range(10):
        t0 = time.perf_counter()
        txt = load(lib, path)
        runs_ms.append((time.perf_counter() - t0) * 1000)

    print(json.dumps({
        "best_ms":   min(runs_ms),
        "median_ms": sorted(runs_ms)[len(runs_ms) // 2],
        "mean_ms":   sum(runs_ms) / len(runs_ms),
        "chars":     len(txt),
    }))


if __name__ == "__main__":
    main()
